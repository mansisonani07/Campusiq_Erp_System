from datetime import date, datetime
from urllib.parse import quote
from fastapi import APIRouter, Depends, Form, Request, HTTPException, status
from fastapi.responses import RedirectResponse, PlainTextResponse, JSONResponse
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session
from ..database import get_db
from .. import models
from ..auth import hash_password
from ..services import risk_score, recommendation_from_risk, dashboard_metrics, fee_reminders, log_action
from ..permissions import require_permission, has_permission, ROLE_PERMISSIONS
from ..form_parser import get_form_value

router = APIRouter(prefix="/admin", tags=["admin"])


def require_admin(request: Request):
    if not request.session.get("role"):
        raise HTTPException(status_code=403, detail="Login required")


def require_perm(permission: str):
    def checker(request: Request):
        require_admin(request)
        require_permission(request, permission)
    return checker


def can_manage(request: Request, category: str) -> bool:
    role = request.session.get("role")
    return has_permission(role, f"{category}.manage")


# ==================== MAIN DASHBOARD ====================

@router.get("/dashboard")
def dashboard(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("dashboard.view"))):
    metrics = dashboard_metrics(db)
    user_id = request.session.get("user_id")
    user = db.query(models.User).filter(models.User.id == user_id).first()
    role = request.session.get("role", "admin")
    announcements = (
        db.query(models.Announcement)
        .filter(models.Announcement.audience.in_(["all", "staff", "admin"]))
        .order_by(models.Announcement.created_at.desc())
        .limit(6)
        .all()
    )
    return request.app.state.templates.TemplateResponse(
        "admin/dashboard.html",
        {
            "request": request,
            "stats": metrics,
            "user": user,
            "role": role,
            "announcements": announcements,
        },
    )


@router.get("/dashboard/live")
def dashboard_live(db: Session = Depends(get_db), _=Depends(require_perm("dashboard.view"))):
    metrics = dashboard_metrics(db)
    return JSONResponse(
        {
            "students": metrics["students"],
            "courses": metrics["courses"],
            "pending_fees": metrics["pending_fees"],
            "collection_rate": metrics["collection_rate"],
            "total_due": metrics["total_due"],
            "total_paid": metrics["total_paid"],
            "pending_amount": metrics["pending_amount"],
            "attendance_30d": metrics["attendance_30d"],
            "attendance_trend": metrics["attendance_trend"],
            "announcements_30d": metrics["announcements_30d"],
            "announcements_trend": metrics["announcements_trend"],
            "high_risk_30d": metrics["high_risk_30d"],
            "high_risk_trend": metrics["high_risk_trend"],
            "overdue_fees": metrics["overdue_fees"],
            "due_soon_fees": metrics["due_soon_fees"],
            "due_today_fees": metrics["due_today_fees"],
            "attendance_today": metrics["attendance_today"],
            "present_today": metrics["present_today"],
            "absent_today": metrics["absent_today"],
            "attendance_today_pct": metrics["attendance_today_pct"],
            "staff_count": metrics.get("staff_count", 0),
            "academic_year": metrics["academic_year"],
            "semester": metrics["semester"],
            "upcoming_assessments": metrics.get("upcoming_assessments", []),
            "action_center": metrics.get("action_center", []),
            "generated_at": datetime.utcnow().isoformat() + "Z",
        }
    )


# ==================== STUDENT MANAGEMENT ====================

@router.get("/students")
def students_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("students.view"))):
    students = db.query(models.StudentProfile).all()
    branches = db.query(models.Branch).all()
    semesters = db.query(models.Semester).all()
    return request.app.state.templates.TemplateResponse(
        "admin/students.html",
        {"request": request, "students": students, "branches": branches, "semesters": semesters, "error": None, "can_manage": can_manage(request, "students")},
    )


@router.get("/students/{student_id}/edit")
def edit_student_page(student_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("students.view"))):
    profile = db.query(models.StudentProfile).filter(models.StudentProfile.id == student_id).first()
    if not profile:
        raise HTTPException(status_code=404, detail="Student not found")
    
    user = db.query(models.User).filter(models.User.id == profile.user_id).first()
    branches = db.query(models.Branch).all()
    semesters = db.query(models.Semester).all()
    
    return request.app.state.templates.TemplateResponse(
        "admin/edit_student.html",
        {
            "request": request,
            "student": profile,
            "user": user,
            "branches": branches,
            "semesters": semesters,
            "can_manage": can_manage(request, "students")
        },
    )


@router.post("/students/{student_id}/edit")
def edit_student(
    student_id: int,
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_perm("students.manage")),
):
    profile = db.query(models.StudentProfile).filter(models.StudentProfile.id == student_id).first()
    if not profile:
        raise HTTPException(status_code=404, detail="Student not found")
    
    user = db.query(models.User).filter(models.User.id == profile.user_id).first()
    
    name = get_form_value(request, "name")
    email = get_form_value(request, "email")
    password = get_form_value(request, "password")
    roll_no = get_form_value(request, "roll_no")
    department = get_form_value(request, "department")
    year_str = get_form_value(request, "year")
    year = int(year_str) if year_str and year_str.isdigit() else 1
    phone = get_form_value(request, "phone")
    guardian_name = get_form_value(request, "guardian_name")
    
    # Get branch_id and semester_id from form
    branch_id_str = get_form_value(request, "branch_id")
    semester_id_str = get_form_value(request, "semester_id")
    branch_id = int(branch_id_str) if branch_id_str and branch_id_str.isdigit() else None
    semester_id = int(semester_id_str) if semester_id_str and semester_id_str.isdigit() else None
    
    try:
        # Update user
        user.name = name
        user.email = email
        if password:
            user.password_hash = hash_password(password)
        
        # Update profile
        profile.roll_no = roll_no
        profile.department = department
        profile.year = year
        profile.phone = phone
        profile.guardian_name = guardian_name
        profile.branch_id = branch_id
        profile.semester_id = semester_id
        
        db.commit()
        
        # Log the action
        log_action(db, request, "Student Updated", f"Updated student {name} (Roll: {roll_no})")
        
    except IntegrityError:
        db.rollback()
        branches = db.query(models.Branch).all()
        semesters = db.query(models.Semester).all()
        return request.app.state.templates.TemplateResponse(
            "admin/edit_student.html",
            {
                "request": request,
                "student": profile,
                "user": user,
                "branches": branches,
                "semesters": semesters,
                "error": "Email or roll number already exists.",
                "can_manage": can_manage(request, "students")
            },
            status_code=400,
        )
    return RedirectResponse("/admin/students", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/students")
def add_student(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_perm("students.manage")),
):
    name = get_form_value(request, "name")
    email = get_form_value(request, "email")
    password = get_form_value(request, "password")
    roll_no = get_form_value(request, "roll_no")
    department = get_form_value(request, "department")
    year_str = get_form_value(request, "year")
    year = int(year_str) if year_str and year_str.isdigit() else 1
    phone = get_form_value(request, "phone")
    guardian_name = get_form_value(request, "guardian_name")
    
    # Get branch_id and semester_id from form
    branch_id_str = get_form_value(request, "branch_id")
    semester_id_str = get_form_value(request, "semester_id")
    branch_id = int(branch_id_str) if branch_id_str and branch_id_str.isdigit() else None
    semester_id = int(semester_id_str) if semester_id_str and semester_id_str.isdigit() else None
    
    try:
        user = models.User(name=name, email=email, password_hash=hash_password(password), role="student")
        db.add(user)
        db.flush()
        profile = models.StudentProfile(
            user_id=user.id,
            roll_no=roll_no,
            department=department,
            year=year,
            phone=phone,
            guardian_name=guardian_name,
            branch_id=branch_id,
            semester_id=semester_id,
        )
        db.add(profile)
        db.commit()
        
        # Log the action
        log_action(db, request, "Student Added", f"Added student {name} (Roll: {roll_no})")
        
    except IntegrityError:
        db.rollback()
        students = db.query(models.StudentProfile).all()
        branches = db.query(models.Branch).all()
        semesters = db.query(models.Semester).all()
        return request.app.state.templates.TemplateResponse(
            "admin/students.html",
            {"request": request, "students": students, "branches": branches, "semesters": semesters, "error": "Email or roll number already exists.", "can_manage": can_manage(request, "students")},
            status_code=400,
        )
    return RedirectResponse("/admin/students", status_code=status.HTTP_303_SEE_OTHER)


# ==================== COURSES ====================

@router.get("/courses")
def courses_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("courses.view"))):
    courses = db.query(models.Course).all()
    return request.app.state.templates.TemplateResponse(
        "admin/courses.html", {"request": request, "courses": courses, "can_manage": can_manage(request, "courses")}
    )


@router.post("/courses")
def add_course(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_perm("courses.manage")),
):
    code = get_form_value(request, "code")
    title = get_form_value(request, "title")
    credits_str = get_form_value(request, "credits")
    credits = int(credits_str) if credits_str and credits_str.isdigit() else 0
    semester = get_form_value(request, "semester")
    db.add(models.Course(code=code, title=title, credits=credits, semester=semester))
    try:
        db.commit()
        
        # Log the action
        log_action(db, request, "Course Added", f"Added course {code} - {title}")
        
    except IntegrityError:
        db.rollback()
    return RedirectResponse("/admin/courses", status_code=status.HTTP_303_SEE_OTHER)


# ==================== ENROLLMENTS ====================

@router.get("/enrollments")
def enrollments_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("enrollments.view"))):
    enrollments = db.query(models.Enrollment).all()
    students = db.query(models.StudentProfile).all()
    courses = db.query(models.Course).all()
    return request.app.state.templates.TemplateResponse(
        "admin/enrollments.html",
        {
            "request": request,
            "enrollments": enrollments,
            "students": students,
            "courses": courses,
            "can_manage": can_manage(request, "enrollments"),
        },
    )


@router.post("/enrollments")
def add_enrollment(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_perm("enrollments.manage")),
):
    student_id_str = get_form_value(request, "student_id")
    course_id_str = get_form_value(request, "course_id")
    student_id = int(student_id_str) if student_id_str and student_id_str.isdigit() else 0
    course_id = int(course_id_str) if course_id_str and course_id_str.isdigit() else 0
    db.add(models.Enrollment(student_id=student_id, course_id=course_id))
    try:
        db.commit()
        
        # Log the action
        log_action(db, request, "Enrollment Added", f"Enrolled student {student_id} in course {course_id}")
        
    except IntegrityError:
        db.rollback()
    return RedirectResponse("/admin/enrollments", status_code=status.HTTP_303_SEE_OTHER)


# ==================== ATTENDANCE ====================

@router.get("/attendance")
def attendance_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("attendance.view"))):
    # Get filter parameters
    branch_id = request.query_params.get("branch_id")
    semester_id = request.query_params.get("semester_id")
    
    # Get branches and semesters for dropdowns
    branches = db.query(models.Branch).all()
    semesters = db.query(models.Semester).all()
    
    # Build query with filters
    query = db.query(models.Attendance)
    
    if branch_id and semester_id:
        # Filter by branch and semester
        query = query.join(models.Enrollment, models.Enrollment.id == models.Attendance.enrollment_id)
        query = query.join(models.StudentProfile, models.StudentProfile.id == models.Enrollment.student_id)
        query = query.filter(
            models.StudentProfile.branch_id == int(branch_id),
            models.StudentProfile.semester_id == int(semester_id)
        )
    
    # Get attendance records
    attendance = query.order_by(models.Attendance.date.desc()).limit(200).all()
    
    # Get enrollments for the form (filtered if needed)
    enrollments_query = db.query(models.Enrollment)
    if branch_id and semester_id:
        enrollments_query = enrollments_query.join(models.StudentProfile, models.StudentProfile.id == models.Enrollment.student_id)
        enrollments_query = enrollments_query.filter(
            models.StudentProfile.branch_id == int(branch_id),
            models.StudentProfile.semester_id == int(semester_id)
        )
    enrollments = enrollments_query.all()
    
    return request.app.state.templates.TemplateResponse(
        "admin/attendance.html",
        {
            "request": request, 
            "attendance": attendance, 
            "enrollments": enrollments,
            "branches": branches,
            "semesters": semesters,
            "selected_branch": int(branch_id) if branch_id else None,
            "selected_semester": int(semester_id) if semester_id else None,
            "can_manage": can_manage(request, "attendance")
        },
    )


@router.post("/attendance")
def add_attendance(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_perm("attendance.manage")),
):
    enrollment_id_str = get_form_value(request, "enrollment_id")
    enrollment_id = int(enrollment_id_str) if enrollment_id_str and enrollment_id_str.isdigit() else 0
    status_text = get_form_value(request, "status_text")
    date_text_str = get_form_value(request, "date_text")
    try:
        date_text = date.fromisoformat(date_text_str) if date_text_str else date.today()
    except ValueError:
        date_text = date.today()
    db.add(models.Attendance(enrollment_id=enrollment_id, status=status_text, date=date_text))
    db.commit()
    return RedirectResponse("/admin/attendance", status_code=status.HTTP_303_SEE_OTHER)


# ==================== GRADES ====================

@router.get("/grades")
def grades_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("grades.view"))):
    grades = db.query(models.Grade).all()
    enrollments = db.query(models.Enrollment).all()
    return request.app.state.templates.TemplateResponse(
        "admin/grades.html",
        {"request": request, "grades": grades, "enrollments": enrollments, "can_manage": can_manage(request, "grades")},
    )


@router.post("/grades")
def add_grade(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_perm("grades.manage")),
):
    enrollment_id_str = get_form_value(request, "enrollment_id")
    enrollment_id = int(enrollment_id_str) if enrollment_id_str and enrollment_id_str.isdigit() else 0
    exam_name = get_form_value(request, "exam_name")
    max_marks_str = get_form_value(request, "max_marks")
    max_marks = float(max_marks_str) if max_marks_str else 0.0
    marks_obtained_str = get_form_value(request, "marks_obtained")
    marks_obtained = float(marks_obtained_str) if marks_obtained_str else 0.0
    db.add(
        models.Grade(
            enrollment_id=enrollment_id,
            exam_name=exam_name,
            max_marks=max_marks,
            marks_obtained=marks_obtained,
        )
    )
    db.commit()
    return RedirectResponse("/admin/grades", status_code=status.HTTP_303_SEE_OTHER)


# ==================== FEES ====================

@router.get("/fees")
def fees_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("fees.view"))):
    fees = db.query(models.FeeRecord).all()
    students = db.query(models.StudentProfile).all()
    reminders = fee_reminders(db)
    notice = request.query_params.get("notice")
    return request.app.state.templates.TemplateResponse(
        "admin/fees.html",
        {
            "request": request,
            "fees": fees,
            "students": students,
            "reminders": reminders,
            "notice": notice,
            "can_manage": can_manage(request, "fees"),
        },
    )


@router.get("/expenses")
def expenses_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("expenses.view"))):
    expenses = db.query(models.Expense).order_by(models.Expense.expense_date.desc()).all()
    return request.app.state.templates.TemplateResponse(
        "admin/expenses.html",
        {
            "request": request,
            "expenses": expenses,
            "can_manage": can_manage(request, "expenses"),
        },
    )


@router.post("/fees/{fee_id}/mark-paid")
def mark_fee_as_paid(fee_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("fees.manage"))):
    fee = db.query(models.FeeRecord).filter(models.FeeRecord.id == fee_id).first()
    if fee:
        fee.status = "paid"
        fee.paid_at = datetime.utcnow()
        fee.amount_paid = fee.amount_due
        db.commit()
        
        # Log the action
        log_action(db, request, "Fee Marked as Paid", f"Marked fee {fee_id} as paid for student {fee.student.roll_no}")
        
    return RedirectResponse("/admin/fees", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/fees")
def add_fee(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_perm("fees.manage")),
):
    student_id_str = get_form_value(request, "student_id")
    student_id = int(student_id_str) if student_id_str and student_id_str.isdigit() else 0
    amount_due_str = get_form_value(request, "amount_due")
    amount_due = float(amount_due_str) if amount_due_str else 0.0
    amount_paid_str = get_form_value(request, "amount_paid", "0")
    amount_paid = float(amount_paid_str) if amount_paid_str else 0.0
    due_date_str = get_form_value(request, "due_date")
    try:
        due_date = date.fromisoformat(due_date_str) if due_date_str else date.today()
    except ValueError:
        due_date = date.today()
    status_text = "paid" if amount_paid >= amount_due else "paid" if amount_paid > 0 else "pending"
    paid_at = datetime.utcnow() if amount_paid > 0 else None
    db.add(
        models.FeeRecord(
            student_id=student_id,
            amount_due=amount_due,
            amount_paid=amount_paid,
            due_date=due_date,
            status=status_text,
            paid_at=paid_at,
        )
    )
    db.commit()
    return RedirectResponse("/admin/fees", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/fees/{fee_id}/remind")
def trigger_fee_reminder(fee_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("fees.manage"))):
    fee = db.query(models.FeeRecord).filter(models.FeeRecord.id == fee_id).first()
    if not fee:
        raise HTTPException(status_code=404, detail="Fee record not found")
    outstanding = max(0.0, fee.amount_due - fee.amount_paid)
    message = (
        f"Reminder queued for {fee.student.user.name} ({fee.student.roll_no}) "
        f"for outstanding amount {outstanding:.2f} due on {fee.due_date.isoformat()}."
    )
    safe_message = quote(message, safe="")
    return RedirectResponse(f"/admin/fees?notice={safe_message}", status_code=status.HTTP_303_SEE_OTHER)


@router.get("/fees/{fee_id}/receipt")
def download_fee_receipt(fee_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("fees.view"))):
    fee = db.query(models.FeeRecord).filter(models.FeeRecord.id == fee_id).first()
    if not fee:
        raise HTTPException(status_code=404, detail="Fee record not found")

    outstanding = max(0.0, fee.amount_due - fee.amount_paid)
    receipt_text = (
        "CampusIQ - Fee Receipt\n"
        "--------------------------------\n"
        f"Receipt Generated: {datetime.utcnow().isoformat()} UTC\n"
        f"Student Name: {fee.student.user.name}\n"
        f"Roll Number: {fee.student.roll_no}\n"
        f"Department: {fee.student.department}\n"
        f"Amount Due: {fee.amount_due:.2f}\n"
        f"Amount Paid: {fee.amount_paid:.2f}\n"
        f"Outstanding: {outstanding:.2f}\n"
        f"Due Date: {fee.due_date.isoformat()}\n"
        f"Status: {fee.status}\n"
    )
    filename = f"fee_receipt_{fee.student.roll_no}_{fee.id}.txt"
    return PlainTextResponse(
        receipt_text,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ==================== ANNOUNCEMENTS ====================

@router.get("/announcements")
def announcement_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("announcements.view"))):
    items = db.query(models.Announcement).order_by(models.Announcement.created_at.desc()).all()
    return request.app.state.templates.TemplateResponse(
        "admin/announcements.html",
        {"request": request, "items": items, "can_manage": can_manage(request, "announcements")},
    )


@router.post("/announcements")
def add_announcement(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_perm("announcements.manage")),
):
    title = get_form_value(request, "title")
    body = get_form_value(request, "body")
    audience = get_form_value(request, "audience", "all")
    db.add(models.Announcement(title=title, body=body, audience=audience))
    db.commit()
    
    # Log the action
    log_action(db, request, "Circular Posted", f"Posted circular: {title}")
    
    return RedirectResponse("/admin/announcements", status_code=status.HTTP_303_SEE_OTHER)


# ==================== SETTINGS ====================

@router.get("/settings")
def settings_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("settings.manage"))):
    settings = {s.key: s.value for s in db.query(models.SiteSetting).all()}
    return request.app.state.templates.TemplateResponse(
        "admin/settings.html", {"request": request, "settings": settings}
    )


@router.post("/settings")
def update_settings(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_perm("settings.manage")),
):
    site_title = get_form_value(request, "site_title")
    site_subtitle = get_form_value(request, "site_subtitle")
    hero_cta = get_form_value(request, "hero_cta")
    pairs = {
        "site_title": site_title,
        "site_subtitle": site_subtitle,
        "hero_cta": hero_cta,
    }
    for key, value in pairs.items():
        setting = db.query(models.SiteSetting).filter_by(key=key).first()
        if setting:
            setting.value = value
        else:
            db.add(models.SiteSetting(key=key, value=value))
    db.commit()
    return RedirectResponse("/admin/settings", status_code=status.HTTP_303_SEE_OTHER)


# ==================== INTERVENTIONS ====================

@router.get("/interventions")
def interventions_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("interventions.view"))):
    rows = db.query(models.Intervention).order_by(models.Intervention.created_at.desc()).all()
    return request.app.state.templates.TemplateResponse(
        "admin/interventions.html",
        {"request": request, "rows": rows, "can_manage": can_manage(request, "interventions")},
    )


@router.post("/interventions/refresh")
def refresh_interventions(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("interventions.manage"))):
    students = db.query(models.StudentProfile).all()
    for student in students:
        score = risk_score(db, student.id)
        rec = recommendation_from_risk(score)
        db.add(models.Intervention(student_id=student.id, risk_score=score, recommendation=rec))
    db.commit()
    return RedirectResponse("/admin/interventions", status_code=status.HTTP_303_SEE_OTHER)


# ==================== USERS ====================

@router.get("/users")
def users_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("users.manage"))):
    users = db.query(models.User).order_by(models.User.role.asc(), models.User.name.asc()).all()
    roles = [r for r in ROLE_PERMISSIONS.keys() if r != "student"]
    return request.app.state.templates.TemplateResponse(
        "admin/users.html", {"request": request, "users": users, "roles": roles}
    )


@router.post("/users/{user_id}/role")
def update_user_role(
    user_id: int,
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_perm("users.manage")),
):
    role = get_form_value(request, "role")
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if role not in ROLE_PERMISSIONS:
        raise HTTPException(status_code=400, detail="Invalid role")
    user.role = role
    db.commit()
    return RedirectResponse("/admin/users", status_code=status.HTTP_303_SEE_OTHER)


# ==================== PHASE 3: BRANCH & SEMESTER MANAGEMENT ====================

@router.get("/branches")
def branches_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("dashboard.view"))):
    branches = db.query(models.Branch).all()
    semesters = db.query(models.Semester).all()
    return request.app.state.templates.TemplateResponse(
        "admin/branches.html",
        {"request": request, "branches": branches, "semesters": semesters},
    )


@router.post("/branches")
def add_branch(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_perm("dashboard.view")),
):
    name = get_form_value(request, "name")
    code = get_form_value(request, "code")
    if name and code:
        db.add(models.Branch(name=name, code=code))
        try:
            db.commit()
            
            # Log the action
            log_action(db, request, "Branch Added", f"Added branch {name} ({code})")
            
        except IntegrityError:
            db.rollback()
    return RedirectResponse("/admin/branches", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/branches/{branch_id}/delete")
def delete_branch(branch_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("dashboard.view"))):
    branch = db.query(models.Branch).filter(models.Branch.id == branch_id).first()
    if branch:
        db.query(models.Semester).filter(models.Semester.branch_id == branch_id).delete()
        db.delete(branch)
        db.commit()
    return RedirectResponse("/admin/branches", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/semesters")
def add_semester(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_perm("dashboard.view")),
):
    branch_id_str = get_form_value(request, "branch_id")
    semester_number_str = get_form_value(request, "semester_number")
    academic_year = get_form_value(request, "academic_year")
    
    branch_id = int(branch_id_str) if branch_id_str and branch_id_str.isdigit() else 0
    semester_number = int(semester_number_str) if semester_number_str and semester_number_str.isdigit() else 0
    
    if branch_id and semester_number and academic_year:
        db.add(models.Semester(branch_id=branch_id, semester_number=semester_number, academic_year=academic_year))
        try:
            db.commit()
        except IntegrityError:
            db.rollback()
    return RedirectResponse("/admin/branches", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/semesters/{semester_id}/delete")
def delete_semester(semester_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("dashboard.view"))):
    semester = db.query(models.Semester).filter(models.Semester.id == semester_id).first()
    if semester:
        db.delete(semester)
        db.commit()
    return RedirectResponse("/admin/branches", status_code=status.HTTP_303_SEE_OTHER)


# ==================== PHASE 3: CIRCULAR MANAGEMENT ====================

@router.get("/circulars")
def circulars_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("announcements.view"))):
    circulars = db.query(models.Circular).order_by(models.Circular.created_at.desc()).all()
    branches = db.query(models.Branch).all()
    semesters = db.query(models.Semester).all()
    return request.app.state.templates.TemplateResponse(
        "admin/circulars.html",
        {"request": request, "circulars": circulars, "branches": branches, "semesters": semesters},
    )


@router.post("/circulars")
def add_circular(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_perm("announcements.manage")),
):
    title = get_form_value(request, "title")
    message = get_form_value(request, "message")
    target_branch = get_form_value(request, "target_branch", "ALL")
    target_semester_str = get_form_value(request, "target_semester", "0")
    target_semester = int(target_semester_str) if target_semester_str and target_semester_str.isdigit() else 0
    
    if title and message:
        db.add(models.Circular(
            title=title,
            message=message,
            target_branch=target_branch if target_branch else "ALL",
            target_semester=target_semester,
        ))
        db.commit()
        
        # Log the action
        log_action(db, request, "Circular Added", f"Added circular: {title}")
        
    return RedirectResponse("/admin/circulars", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/circulars/{circular_id}/delete")
def delete_circular(circular_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("announcements.manage"))):
    circular = db.query(models.Circular).filter(models.Circular.id == circular_id).first()
    if circular:
        db.delete(circular)
        db.commit()
    return RedirectResponse("/admin/circulars", status_code=status.HTTP_303_SEE_OTHER)


# ==================== PHASE 3: TIMETABLE MANAGEMENT ====================

@router.get("/timetables")
def timetables_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("courses.view"))):
    timetables = db.query(models.Timetable).order_by(models.Timetable.uploaded_at.desc()).all()
    branches = db.query(models.Branch).all()
    semesters = db.query(models.Semester).all()
    return request.app.state.templates.TemplateResponse(
        "admin/timetables.html",
        {"request": request, "timetables": timetables, "branches": branches, "semesters": semesters},
    )


@router.post("/timetables")
def add_timetable(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_perm("courses.manage")),
):
    branch_id_str = get_form_value(request, "branch_id")
    semester_id_str = get_form_value(request, "semester_id")
    file_path = get_form_value(request, "file_path", "/static/uploads/timetable.pdf")
    file_name = get_form_value(request, "file_name", "timetable.pdf")
    
    branch_id = int(branch_id_str) if branch_id_str and branch_id_str.isdigit() else 0
    semester_id = int(semester_id_str) if semester_id_str and semester_id_str.isdigit() else 0
    
    if branch_id and semester_id:
        db.add(models.Timetable(
            branch_id=branch_id,
            semester_id=semester_id,
            file_path=file_path,
            file_name=file_name
        ))
        db.commit()
        
        # Log the action
        log_action(db, request, "Timetable Added", f"Added timetable for branch {branch_id}, semester {semester_id}")
        
    return RedirectResponse("/admin/timetables", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/timetables/{timetable_id}/delete")
def delete_timetable(timetable_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("courses.manage"))):
    timetable = db.query(models.Timetable).filter(models.Timetable.id == timetable_id).first()
    if timetable:
        db.delete(timetable)
        db.commit()
    return RedirectResponse("/admin/timetables", status_code=status.HTTP_303_SEE_OTHER)


# ==================== PHASE 3: EXAM SCHEDULE MANAGEMENT ====================

@router.get("/exam-schedules")
def exam_schedules_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("grades.view"))):
    exam_schedules = db.query(models.ExamSchedule).order_by(models.ExamSchedule.exam_date.desc()).all()
    branches = db.query(models.Branch).all()
    semesters = db.query(models.Semester).all()
    return request.app.state.templates.TemplateResponse(
        "admin/exam_schedules.html",
        {"request": request, "exam_schedules": exam_schedules, "branches": branches, "semesters": semesters},
    )


@router.post("/exam-schedules")
def add_exam_schedule(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_perm("grades.manage")),
):
    branch_id_str = get_form_value(request, "branch_id")
    semester_id_str = get_form_value(request, "semester_id")
    subject_name = get_form_value(request, "subject_name")
    exam_date_str = get_form_value(request, "exam_date")
    exam_time = get_form_value(request, "exam_time", "09:00 AM")
    
    branch_id = int(branch_id_str) if branch_id_str and branch_id_str.isdigit() else 0
    semester_id = int(semester_id_str) if semester_id_str and semester_id_str.isdigit() else 0
    
    try:
        exam_date = date.fromisoformat(exam_date_str) if exam_date_str else date.today()
    except ValueError:
        exam_date = date.today()
    
    if branch_id and semester_id and subject_name:
        db.add(models.ExamSchedule(
            branch_id=branch_id,
            semester_id=semester_id,
            subject_name=subject_name,
            exam_date=exam_date,
            exam_time=exam_time
        ))
        db.commit()
        
        # Log the action
        log_action(db, request, "Exam Schedule Added", f"Added exam schedule for {subject_name} in branch {branch_id}, semester {semester_id}")
        
    return RedirectResponse("/admin/exam-schedules", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/exam-schedules/{exam_id}/delete")
def delete_exam_schedule(exam_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("grades.manage"))):
    exam = db.query(models.ExamSchedule).filter(models.ExamSchedule.id == exam_id).first()
    if exam:
        db.delete(exam)
        db.commit()
    return RedirectResponse("/admin/exam-schedules", status_code=status.HTTP_303_SEE_OTHER)


# ==================== PHASE 3: RESULT MANAGEMENT ====================

@router.get("/results")
def results_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("grades.view"))):
    branch_filter = request.query_params.get("branch_id")
    semester_filter = request.query_params.get("semester_id")
    
    results = db.query(models.Result).all()
    if branch_filter:
        results = [r for r in results if r.branch_id == int(branch_filter)]
    if semester_filter:
        results = [r for r in results if r.semester_id == int(semester_filter)]
    
    branches = db.query(models.Branch).all()
    semesters = db.query(models.Semester).all()
    students = db.query(models.StudentProfile).all()
    
    return request.app.state.templates.TemplateResponse(
        "admin/results.html",
        {
            "request": request, 
            "results": results, 
            "branches": branches, 
            "semesters": semesters,
            "students": students,
            "branch_filter": branch_filter,
            "semester_filter": semester_filter
        },
    )


@router.post("/results")
def add_result(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_perm("grades.manage")),
):
    branch_id_str = get_form_value(request, "branch_id")
    semester_id_str = get_form_value(request, "semester_id")
    student_id_str = get_form_value(request, "student_id")
    subject_name = get_form_value(request, "subject_name")
    exam_name = get_form_value(request, "exam_name", "Final")
    max_marks_str = get_form_value(request, "max_marks")
    marks_obtained_str = get_form_value(request, "marks_obtained")
    status_text = get_form_value(request, "status", "draft")
    
    branch_id = int(branch_id_str) if branch_id_str and branch_id_str.isdigit() else 0
    semester_id = int(semester_id_str) if semester_id_str and semester_id_str.isdigit() else 0
    student_id = int(student_id_str) if student_id_str and student_id_str.isdigit() else 0
    max_marks = float(max_marks_str) if max_marks_str else 0.0
    marks_obtained = float(marks_obtained_str) if marks_obtained_str else 0.0
    
    if branch_id and semester_id and student_id and subject_name:
        db.add(models.Result(
            branch_id=branch_id,
            semester_id=semester_id,
            student_id=student_id,
            subject_name=subject_name,
            exam_name=exam_name,
            max_marks=max_marks,
            marks_obtained=marks_obtained,
            status=status_text
        ))
        db.commit()
        
        # Log the action
        log_action(db, request, "Result Added", f"Added result for {subject_name} - {exam_name} for student {student_id}")
        
    return RedirectResponse("/admin/results", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/results/{result_id}/delete")
def delete_result(result_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("grades.manage"))):
    result = db.query(models.Result).filter(models.Result.id == result_id).first()
    if result:
        db.delete(result)
        db.commit()
    return RedirectResponse("/admin/results", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/results/{result_id}/update-status")
def update_result_status(result_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("grades.manage"))):
    status_text = get_form_value(request, "status")
    result = db.query(models.Result).filter(models.Result.id == result_id).first()
    if result and status_text:
        result.status = status_text
        db.commit()
    return RedirectResponse("/admin/results", status_code=status.HTTP_303_SEE_OTHER)


# ==================== PHASE 3: NOTICE BOARD ====================

@router.get("/notices")
def notices_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("announcements.view"))):
    notices = db.query(models.Notice).order_by(models.Notice.created_at.desc()).all()
    branches = db.query(models.Branch).all()
    semesters = db.query(models.Semester).all()
    return request.app.state.templates.TemplateResponse(
        "admin/notices.html",
        {"request": request, "notices": notices, "branches": branches, "semesters": semesters},
    )


@router.post("/notices")
def add_notice(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_perm("announcements.manage")),
):
    title = get_form_value(request, "title")
    content = get_form_value(request, "content")
    target_role = get_form_value(request, "target_role", "all")
    target_branch = get_form_value(request, "target_branch", "ALL")
    target_semester_str = get_form_value(request, "target_semester", "0")
    target_semester = int(target_semester_str) if target_semester_str and target_semester_str.isdigit() else 0
    
    if title and content:
        db.add(models.Notice(
            title=title,
            content=content,
            target_role=target_role,
            target_branch=target_branch if target_branch else "ALL",
            target_semester=target_semester,
        ))
        db.commit()
        
        # Log the action
        log_action(db, request, "Notice Added", f"Added notice: {title}")
        
    return RedirectResponse("/admin/notices", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/notices/{notice_id}/delete")
def delete_notice(notice_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("announcements.manage"))):
    notice = db.query(models.Notice).filter(models.Notice.id == notice_id).first()
    if notice:
        db.delete(notice)
        db.commit()
    return RedirectResponse("/admin/notices", status_code=status.HTTP_303_SEE_OTHER)


# ==================== PHASE 3: STUDENT MANAGEMENT NEW ====================

@router.get("/students-new")
def students_new_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("students.view"))):
    students = db.query(models.StudentProfile).all()
    branches = db.query(models.Branch).all()
    semesters = db.query(models.Semester).all()
    
    filter_branch = request.query_params.get("branch", "")
    filter_semester = request.query_params.get("semester", "")
    
    filtered_students = students
    if filter_branch:
        filtered_students = [s for s in students if s.department == filter_branch]
    if filter_semester:
        filtered_students = [s for s in filtered_students if str(s.year) == filter_semester]
    
    return request.app.state.templates.TemplateResponse(
        "admin/students_new.html",
        {
            "request": request, 
            "students": filtered_students,
            "branches": branches,
            "semesters": semesters,
            "filter_branch": filter_branch,
            "filter_semester": filter_semester,
        },
    )


@router.post("/students-new")
def add_student_new(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_perm("students.manage")),
):
    name = get_form_value(request, "name")
    email = get_form_value(request, "email")
    password = get_form_value(request, "password")
    roll_no = get_form_value(request, "roll_no")
    branch = get_form_value(request, "branch")
    semester_str = get_form_value(request, "semester")
    phone = get_form_value(request, "phone")
    address = get_form_value(request, "address")
    
    year = int(semester_str) if semester_str and semester_str.isdigit() else 1
    
    try:
        user = models.User(name=name, email=email, password_hash=hash_password(password), role="student")
        db.add(user)
        db.flush()
        profile = models.StudentProfile(
            user_id=user.id,
            roll_no=roll_no,
            department=branch,
            year=year,
            phone=phone,
            guardian_name="",
        )
        db.add(profile)
        db.commit()
        
        # Log the action
        log_action(db, request, "Student Added", f"Added student {name} (Roll: {roll_no})")
        
    except IntegrityError:
        db.rollback()
    
    return RedirectResponse("/admin/students-new", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/students-new/{student_id}/delete")
def delete_student_new(student_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("students.manage"))):
    profile = db.query(models.StudentProfile).filter(models.StudentProfile.id == student_id).first()
    if profile:
        user = db.query(models.User).filter(models.User.id == profile.user_id).first()
        if user:
            db.delete(user)
        db.delete(profile)
        db.commit()
    return RedirectResponse("/admin/students-new", status_code=status.HTTP_303_SEE_OTHER)


# ==================== PHASE 3: FEE MANAGEMENT NEW ====================

@router.get("/fees-new")
def fees_new_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("fees.view"))):
    fees = db.query(models.FeeRecord).all()
    students = db.query(models.StudentProfile).all()
    branches = db.query(models.Branch).all()
    semesters = db.query(models.Semester).all()
    
    filter_branch = request.query_params.get("branch", "")
    filter_semester = request.query_params.get("semester", "")
    
    return request.app.state.templates.TemplateResponse(
        "admin/fees_new.html",
        {
            "request": request, 
            "fees": fees,
            "students": students,
            "branches": branches,
            "semesters": semesters,
            "filter_branch": filter_branch,
            "filter_semester": filter_semester,
        },
    )


@router.post("/fees-new")
def add_fee_new(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_perm("fees.manage")),
):
    student_id_str = get_form_value(request, "student_id")
    amount_due_str = get_form_value(request, "amount_due")
    amount_paid_str = get_form_value(request, "amount_paid", "0")
    due_date_str = get_form_value(request, "due_date")
    
    student_id = int(student_id_str) if student_id_str and student_id_str.isdigit() else 0
    amount_due = float(amount_due_str) if amount_due_str else 0.0
    amount_paid = float(amount_paid_str) if amount_paid_str else 0.0
    
    try:
        due_date = date.fromisoformat(due_date_str) if due_date_str else date.today()
    except ValueError:
        due_date = date.today()
    
    status_text = "paid" if amount_paid >= amount_due else "pending"
    paid_at = datetime.utcnow() if amount_paid > 0 else None
    
    if student_id and amount_due:
        db.add(models.FeeRecord(
            student_id=student_id,
            amount_due=amount_due,
            amount_paid=amount_paid,
            due_date=due_date,
            status=status_text,
            paid_at=paid_at,
        ))
        db.commit()
    
    return RedirectResponse("/admin/fees-new", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/fees-new/{fee_id}/mark-paid")
def mark_fee_paid(fee_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("fees.manage"))):
    fee = db.query(models.FeeRecord).filter(models.FeeRecord.id == fee_id).first()
    if fee:
        fee.status = "paid"
        fee.paid_at = datetime.utcnow()
        fee.amount_paid = fee.amount_due
        db.commit()
    return RedirectResponse("/admin/fees-new", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/fees-new/{fee_id}/delete")
def delete_fee_new(fee_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("fees.manage"))):
    fee = db.query(models.FeeRecord).filter(models.FeeRecord.id == fee_id).first()
    if fee:
        db.delete(fee)
        db.commit()
    return RedirectResponse("/admin/fees-new", status_code=status.HTTP_303_SEE_OTHER)


# ==================== TRANSPORT MANAGEMENT ====================

@router.get("/transport")
def transport_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("transport.view"))):
    routes = db.query(models.TransportRoute).all()
    return request.app.state.templates.TemplateResponse(
        "admin/transport.html",
        {"request": request, "routes": routes, "title": "Transport Management"},
    )


@router.post("/transport/add")
def add_transport_route(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("transport.manage"))):
    route_name = get_form_value(request, "route_name")
    bus_number = get_form_value(request, "bus_number")
    departure_time = get_form_value(request, "departure_time")
    stops = get_form_value(request, "stops")
    
    if route_name and bus_number:
        db.add(models.TransportRoute(
            route_name=route_name,
            bus_number=bus_number,
            departure_time=departure_time,
            stops=stops
        ))
        db.commit()
    
    return RedirectResponse("/admin/transport", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/transport/{route_id}/delete")
def delete_transport_route(route_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("transport.manage"))):
    route = db.query(models.TransportRoute).filter(models.TransportRoute.id == route_id).first()
    if route:
        # Delete student assignments first
        db.query(models.StudentTransport).filter(models.StudentTransport.route_id == route_id).delete()
        db.delete(route)
        db.commit()
    
    return RedirectResponse("/admin/transport", status_code=status.HTTP_303_SEE_OTHER)


@router.get("/transport/assign")
def assign_transport_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("transport.view"))):
    routes = db.query(models.TransportRoute).all()
    students = db.query(models.StudentProfile).all()
    assignments = db.query(models.StudentTransport).all()
    return request.app.state.templates.TemplateResponse(
        "admin/transport_assign.html",
        {"request": request, "routes": routes, "students": students, "assignments": assignments, "title": "Assign Transport"},
    )


@router.post("/transport/assign")
def assign_transport(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("transport.manage"))):
    student_id = get_form_value(request, "student_id")
    route_id = get_form_value(request, "route_id")
    pickup_stop = get_form_value(request, "pickup_stop")
    
    if student_id and route_id:
        # Remove existing assignment
        db.query(models.StudentTransport).filter(models.StudentTransport.student_id == student_id).delete()
        
        db.add(models.StudentTransport(
            student_id=int(student_id),
            route_id=int(route_id),
            pickup_stop=pickup_stop
        ))
        db.commit()
    
    return RedirectResponse("/admin/transport/assign", status_code=status.HTTP_303_SEE_OTHER)


# ==================== HOSTEL MANAGEMENT ====================

@router.get("/hostel")
def hostel_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("hostel.view"))):
    rooms = db.query(models.HostelRoom).all()
    return request.app.state.templates.TemplateResponse(
        "admin/hostel.html",
        {"request": request, "rooms": rooms, "title": "Hostel Management"},
    )


@router.post("/hostel/add")
def add_hostel_room(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("hostel.manage"))):
    room_number = get_form_value(request, "room_number")
    floor = get_form_value(request, "floor")
    capacity = get_form_value(request, "capacity", "4")
    
    if room_number:
        db.add(models.HostelRoom(
            room_number=room_number,
            floor=int(floor) if floor else 1,
            capacity=int(capacity) if capacity else 4
        ))
        db.commit()
    
    return RedirectResponse("/admin/hostel", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/hostel/{room_id}/delete")
def delete_hostel_room(room_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("hostel.manage"))):
    room = db.query(models.HostelRoom).filter(models.HostelRoom.id == room_id).first()
    if room:
        # Delete student assignments first
        db.query(models.StudentHostel).filter(models.StudentHostel.room_id == room_id).delete()
        db.delete(room)
        db.commit()
    
    return RedirectResponse("/admin/hostel", status_code=status.HTTP_303_SEE_OTHER)


@router.get("/hostel/assign")
def assign_hostel_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("hostel.view"))):
    rooms = db.query(models.HostelRoom).all()
    students = db.query(models.StudentProfile).all()
    assignments = db.query(models.StudentHostel).all()
    return request.app.state.templates.TemplateResponse(
        "admin/hostel_assign.html",
        {"request": request, "rooms": rooms, "students": students, "assignments": assignments, "title": "Assign Hostel"},
    )


@router.post("/hostel/assign")
def assign_hostel(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("hostel.manage"))):
    student_id = get_form_value(request, "student_id")
    room_id = get_form_value(request, "room_id")
    bed_number = get_form_value(request, "bed_number")
    
    if student_id and room_id:
        # Remove existing assignment
        db.query(models.StudentHostel).filter(models.StudentHostel.student_id == student_id).delete()
        
        db.add(models.StudentHostel(
            student_id=int(student_id),
            room_id=int(room_id),
            bed_number=bed_number
        ))
        
        # Update room occupancy
        room = db.query(models.HostelRoom).filter(models.HostelRoom.id == int(room_id)).first()
        if room:
            room.current_occupancy = (room.current_occupancy or 0) + 1
        
        db.commit()
    
    return RedirectResponse("/admin/hostel/assign", status_code=status.HTTP_303_SEE_OTHER)


# ==================== SUBJECT MANAGEMENT ====================

@router.get("/subjects")
def subjects_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("courses.view"))):
    subjects = db.query(models.Subject).all()
    branches = db.query(models.Branch).all()
    semesters = db.query(models.Semester).all()
    return request.app.state.templates.TemplateResponse(
        "admin/subjects.html",
        {"request": request, "subjects": subjects, "branches": branches, "semesters": semesters, "title": "Subject Management"},
    )


@router.post("/subjects/add")
def add_subject(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("courses.manage"))):
    name = get_form_value(request, "name")
    code = get_form_value(request, "code")
    branch_id = get_form_value(request, "branch_id")
    semester_id = get_form_value(request, "semester_id")
    
    if name and code:
        db.add(models.Subject(
            name=name,
            code=code,
            branch_id=int(branch_id) if branch_id else None,
            semester_id=int(semester_id) if semester_id else None
        ))
        db.commit()
        
        # Log the action
        log_action(db, request, "Subject Added", f"Added subject {code} - {name}")
        
    return RedirectResponse("/admin/subjects", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/subjects/{subject_id}/delete")
def delete_subject(subject_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("courses.manage"))):
    subject = db.query(models.Subject).filter(models.Subject.id == subject_id).first()
    if subject:
        db.delete(subject)
        db.commit()
        
        # Log the action
        log_action(db, request, "Subject Deleted", f"Deleted subject {subject.code} - {subject.name}")
        
    return RedirectResponse("/admin/subjects", status_code=status.HTTP_303_SEE_OTHER)


# ==================== FEE STRUCTURE MANAGEMENT ====================

@router.get("/fee-structure")
def fee_structure_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("fees.view"))):
    fee_structures = db.query(models.FeeStructure).all()
    branches = db.query(models.Branch).all()
    semesters = db.query(models.Semester).all()
    return request.app.state.templates.TemplateResponse(
        "admin/fee_structure.html",
        {"request": request, "fee_structures": fee_structures, "branches": branches, "semesters": semesters, "title": "Fee Structure Management"},
    )


@router.post("/fee-structure/add")
def add_fee_structure(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("fees.manage"))):
    from datetime import date
    branch_id = get_form_value(request, "branch_id")
    semester_id = get_form_value(request, "semester_id")
    total_amount = get_form_value(request, "total_amount")
    academic_year = get_form_value(request, "academic_year")
    
    if branch_id and semester_id and total_amount and academic_year:
        # Create fee structure
        fs = db.add(models.FeeStructure(
            branch_id=int(branch_id),
            semester_id=int(semester_id),
            total_amount=float(total_amount),
            academic_year=academic_year
        ))
        db.flush()
        
        # Auto-create fee records for all students in branch+semester
        students = db.query(models.StudentProfile).filter(
            models.StudentProfile.branch_id == int(branch_id),
            models.StudentProfile.semester_id == int(semester_id)
        ).all()
        
        for student in students:
            # Check if fee record already exists
            existing = db.query(models.FeeRecord).filter(
                models.FeeRecord.student_id == student.id
            ).first()
            if not existing:
                db.add(models.FeeRecord(
                    student_id=student.id,
                    amount_due=float(total_amount),
                    amount_paid=0,
                    due_date=date(2026, 6, 30),
                    status="pending"
                ))
        
        db.commit()
    
    return RedirectResponse("/admin/fee-structure", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/fee-structure/{fs_id}/delete")
def delete_fee_structure(fs_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("fees.manage"))):
    fs = db.query(models.FeeStructure).filter(models.FeeStructure.id == fs_id).first()
    if fs:
        db.delete(fs)
        db.commit()
    return RedirectResponse("/admin/fee-structure", status_code=status.HTTP_303_SEE_OTHER)


# ==================== STUDENT ID CARD ====================

@router.get("/students/{student_id}/id-card")
def student_id_card(student_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("students.view"))):
    profile = db.query(models.StudentProfile).filter(models.StudentProfile.id == student_id).first()
    if not profile:
        raise HTTPException(status_code=404, detail="Student not found")
    
    user = db.query(models.User).filter(models.User.id == profile.user_id).first()
    branch = db.query(models.Branch).filter(models.Branch.id == profile.branch_id).first() if profile.branch_id else None
    semester = db.query(models.Semester).filter(models.Semester.id == profile.semester_id).first() if profile.semester_id else None
    
    return request.app.state.templates.TemplateResponse(
        "admin/student_id_card.html",
        {
            "request": request,
            "student": user,
            "profile": profile,
            "branch": branch,
            "semester": semester,
            "title": f"ID Card - {user.name if user else 'Student'}",
        },
    )


# ==================== ATTENDANCE REPORT ====================

@router.get("/attendance/report")
def attendance_report_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("attendance.view"))):
    branches = db.query(models.Branch).all()
    semesters = db.query(models.Semester).all()
    
    branch_id = request.query_params.get("branch_id")
    semester_id = request.query_params.get("semester_id")
    
    report_data = []
    if branch_id and semester_id:
        students = db.query(models.StudentProfile).filter(
            models.StudentProfile.branch_id == int(branch_id),
            models.StudentProfile.semester_id == int(semester_id)
        ).all()
        
        for student in students:
            user = db.query(models.User).filter(models.User.id == student.user_id).first()
            enrollments = db.query(models.Enrollment).filter(
                models.Enrollment.student_id == student.id
            ).all()
            
            total_classes = 0
            present = 0
            absent = 0
            
            for enrollment in enrollments:
                attendance = db.query(models.Attendance).filter(
                    models.Attendance.enrollment_id == enrollment.id
                ).all()
                total_classes += len(attendance)
                present += sum(1 for a in attendance if a.status == "present")
                absent += sum(1 for a in attendance if a.status == "absent")
            
            pct = round((present / total_classes * 100), 2) if total_classes > 0 else 0
            report_data.append({
                "name": user.name if user else "Unknown",
                "roll_no": student.roll_no,
                "total": total_classes,
                "present": present,
                "absent": absent,
                "percentage": pct
            })
    
    return request.app.state.templates.TemplateResponse(
        "admin/attendance_report.html",
        {
            "request": request,
            "branches": branches,
            "semesters": semesters,
            "report_data": report_data,
            "selected_branch": int(branch_id) if branch_id else None,
            "selected_semester": int(semester_id) if semester_id else None,
            "title": "Attendance Report",
        },
    )


# ==================== BULK RESULT STATUS ====================

@router.post("/results/bulk-status")
def bulk_result_status(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("results.manage"))):
    branch_id = get_form_value(request, "branch_id")
    semester_id = get_form_value(request, "semester_id")
    action = get_form_value(request, "action")  # publish or coming_soon
    
    if branch_id and semester_id and action:
        new_status = "published" if action == "publish" else "coming_soon"
        
        db.query(models.Result).filter(
            models.Result.branch_id == int(branch_id),
            models.Result.semester_id == int(semester_id)
        ).update({"status": new_status})
        db.commit()
    
    return RedirectResponse("/admin/results", status_code=status.HTTP_303_SEE_OTHER)


# ==================== STUDENT PROMOTION ====================

@router.get("/promote-students")
def promote_students_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("students.manage"))):
    branches = db.query(models.Branch).all()
    semesters = db.query(models.Semester).all()
    return request.app.state.templates.TemplateResponse(
        "admin/promote_students.html",
        {"request": request, "branches": branches, "semesters": semesters, "title": "Student Promotion"},
    )


@router.post("/promote-students/preview")
def promote_students_preview(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("students.manage"))):
    branch_id = get_form_value(request, "branch_id")
    current_sem = get_form_value(request, "current_semester")
    new_sem = get_form_value(request, "new_semester")
    
    students = []
    if branch_id and current_sem and new_sem:
        students = db.query(models.StudentProfile).filter(
            models.StudentProfile.branch_id == int(branch_id),
            models.StudentProfile.semester_id == int(current_sem)
        ).all()
        
        for s in students:
            user = db.query(models.User).filter(models.User.id == s.user_id).first()
            s.user_name = user.name if user else "Unknown"
    
    branches = db.query(models.Branch).all()
    semesters = db.query(models.Semester).all()
    
    return request.app.state.templates.TemplateResponse(
        "admin/promote_students.html",
        {
            "request": request,
            "branches": branches,
            "semesters": semesters,
            "preview_students": students,
            "selected_branch": int(branch_id) if branch_id else None,
            "selected_current": int(current_sem) if current_sem else None,
            "selected_new": int(new_sem) if new_sem else None,
            "title": "Student Promotion",
        },
    )


@router.post("/promote-students/confirm")
def promote_students_confirm(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("students.manage"))):
    branch_id = get_form_value(request, "branch_id")
    current_sem = get_form_value(request, "current_semester")
    new_sem = get_form_value(request, "new_semester")
    
    if branch_id and current_sem and new_sem:
        count = db.query(models.StudentProfile).filter(
            models.StudentProfile.branch_id == int(branch_id),
            models.StudentProfile.semester_id == int(current_sem)
        ).update({"semester_id": int(new_sem)})
        db.commit()
        message = f"{count} students promoted successfully"
    else:
        message = "Please select all fields"
    
    branches = db.query(models.Branch).all()
    semesters = db.query(models.Semester).all()
    
    return request.app.state.templates.TemplateResponse(
        "admin/promote_students.html",
        {
            "request": request,
            "branches": branches,
            "semesters": semesters,
            "message": message,
            "title": "Student Promotion",
        },
    )


# ==================== ACADEMIC YEARS ====================

@router.get("/academic-years")
def academic_years_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("settings.manage"))):
    academic_years = db.query(models.AcademicYear).order_by(models.AcademicYear.start_date.desc()).all()
    return request.app.state.templates.TemplateResponse(
        "admin/academic_years.html",
        {"request": request, "academic_years": academic_years, "title": "Academic Years"},
    )


@router.post("/academic-years/add")
def add_academic_year(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("settings.manage"))):
    from datetime import date
    name = get_form_value(request, "name")
    start_date_str = get_form_value(request, "start_date")
    end_date_str = get_form_value(request, "end_date")
    
    if name and start_date_str and end_date_str:
        try:
            start_date = date.fromisoformat(start_date_str)
            end_date = date.fromisoformat(end_date_str)
            db.add(models.AcademicYear(name=name, start_date=start_date, end_date=end_date))
            db.commit()
        except:
            pass
    
    return RedirectResponse("/admin/academic-years", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/academic-years/{year_id}/set-active")
def set_academic_year_active(year_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("settings.manage"))):
    # Deactivate all
    db.query(models.AcademicYear).update({"is_active": 0})
    # Activate selected
    year = db.query(models.AcademicYear).filter(models.AcademicYear.id == year_id).first()
    if year:
        year.is_active = 1
        db.commit()
    
    return RedirectResponse("/admin/academic-years", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/academic-years/{year_id}/delete")
def delete_academic_year(year_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("settings.manage"))):
    year = db.query(models.AcademicYear).filter(models.AcademicYear.id == year_id).first()
    if year:
        db.delete(year)
        db.commit()
    return RedirectResponse("/admin/academic-years", status_code=status.HTTP_303_SEE_OTHER)


# ==================== AUDIT LOG ====================

@router.get("/audit-log")
def audit_log_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("dashboard.view"))):
    logs = db.query(models.AuditLog).order_by(models.AuditLog.created_at.desc()).limit(100).all()
    return request.app.state.templates.TemplateResponse(
        "admin/audit_log.html",
        {"request": request, "logs": logs, "title": "Audit Log"},
    )


# ==================== STUDY MATERIALS (ADMIN) ====================

@router.get("/study-materials")
def study_materials_admin_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("courses.view"))):
    materials = db.query(models.StudyMaterial).order_by(models.StudyMaterial.created_at.desc()).all()
    branches = db.query(models.Branch).all()
    semesters = db.query(models.Semester).all()
    subjects = db.query(models.Subject).all()
    return request.app.state.templates.TemplateResponse(
        "admin/study_materials.html",
        {"request": request, "materials": materials, "branches": branches, "semesters": semesters, "subjects": subjects, "title": "Study Materials"},
    )


@router.post("/study-materials/add")
def add_study_material(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("courses.manage"))):
    title = get_form_value(request, "title")
    description = get_form_value(request, "description")
    branch_id_str = get_form_value(request, "branch_id")
    semester_id_str = get_form_value(request, "semester_id")
    subject_id_str = get_form_value(request, "subject_id")
    
    branch_id = int(branch_id_str) if branch_id_str and branch_id_str.isdigit() else None
    semester_id = int(semester_id_str) if semester_id_str and semester_id_str.isdigit() else None
    subject_id = int(subject_id_str) if subject_id_str and subject_id_str.isdigit() else None
    
    if title:
        db.add(models.StudyMaterial(
            title=title,
            description=description,
            file_path="/static/uploads/study-materials/",
            branch_id=branch_id,
            semester_id=semester_id,
            subject_id=subject_id
        ))
        db.commit()
    
    return RedirectResponse("/admin/study-materials", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/study-materials/{material_id}/delete")
def delete_study_material(material_id: int, request: Request, db: Session = Depends(get_db), _=Depends(require_perm("courses.manage"))):
    material = db.query(models.StudyMaterial).filter(models.StudyMaterial.id == material_id).first()
    if material:
        db.delete(material)
        db.commit()
    return RedirectResponse("/admin/study-materials", status_code=status.HTTP_303_SEE_OTHER)


# ==================== EXPORT TO EXCEL ====================

@router.get("/students/export")
def export_students(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("students.view"))):
    """Export students to Excel"""
    import io
    try:
        import openpyxl
    except ImportError:
        return PlainTextResponse("openpyxl not installed", status_code=500)
    
    students = db.query(models.StudentProfile).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    
    # Headers
    ws.append(["Roll No", "Name", "Email", "Branch", "Semester", "Phone", "Fee Status"])
    
    for student in students:
        user = db.query(models.User).filter(models.User.id == student.user_id).first()
        branch = db.query(models.Branch).filter(models.Branch.id == student.branch_id).first() if student.branch_id else None
        semester = db.query(models.Semester).filter(models.Semester.id == student.semester_id).first() if student.semester_id else None
        
        # Get fee status
        fee = db.query(models.FeeRecord).filter(models.FeeRecord.student_id == student.id).first()
        fee_status = fee.status if fee else "No Fee Record"
        
        ws.append([
            student.roll_no,
            user.name if user else "",
            user.email if user else "",
            branch.name if branch else "",
            semester.semester_number if semester else "",
            student.phone or "",
            fee_status
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=students.xlsx"}
    )


@router.get("/fees/export")
def export_fees(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("fees.view"))):
    """Export fees to Excel"""
    import io
    try:
        import openpyxl
    except ImportError:
        return PlainTextResponse("openpyxl not installed", status_code=500)
    
    fees = db.query(models.FeeRecord).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fees"
    
    # Headers
    ws.append(["Student Name", "Roll No", "Branch", "Semester", "Amount Due", "Amount Paid", "Pending", "Status"])
    
    for fee in fees:
        student = db.query(models.StudentProfile).filter(models.StudentProfile.id == fee.student_id).first()
        if student:
            user = db.query(models.User).filter(models.User.id == student.user_id).first()
            branch = db.query(models.Branch).filter(models.Branch.id == student.branch_id).first() if student.branch_id else None
            semester = db.query(models.Semester).filter(models.Semester.id == student.semester_id).first() if student.semester_id else None
            
            pending = fee.amount_due - fee.amount_paid
            
            ws.append([
                user.name if user else "",
                student.roll_no,
                branch.name if branch else "",
                semester.semester_number if semester else "",
                fee.amount_due,
                fee.amount_paid,
                pending,
                fee.status
            ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=fees.xlsx"}
    )


@router.get("/attendance/report/export")
def export_attendance_report(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("attendance.view"))):
    """Export attendance report to Excel"""
    import io
    try:
        import openpyxl
    except ImportError:
        return PlainTextResponse("openpyxl not installed", status_code=500)
    
    branch_id = request.query_params.get("branch_id")
    semester_id = request.query_params.get("semester_id")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Headers
    ws.append(["Roll Number", "Student Name", "Total Classes", "Present", "Absent", "Percentage"])
    
    if branch_id and semester_id:
        students = db.query(models.StudentProfile).filter(
            models.StudentProfile.branch_id == int(branch_id),
            models.StudentProfile.semester_id == int(semester_id)
        ).all()
        
        for student in students:
            user = db.query(models.User).filter(models.User.id == student.user_id).first()
            enrollments = db.query(models.Enrollment).filter(
                models.Enrollment.student_id == student.id
            ).all()
            
            total_classes = 0
            present = 0
            absent = 0
            
            for enrollment in enrollments:
                attendance = db.query(models.Attendance).filter(
                    models.Attendance.enrollment_id == enrollment.id
                ).all()
                total_classes += len(attendance)
                present += sum(1 for a in attendance if a.status == "present")
                absent += sum(1 for a in attendance if a.status == "absent")
            
            pct = round((present / total_classes * 100), 2) if total_classes > 0 else 0
            
            ws.append([
                student.roll_no,
                user.name if user else "",
                total_classes,
                present,
                absent,
                f"{pct}%"
            ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=attendance_report.xlsx"}
    )


# ==================== BULK STUDENT IMPORT ====================

@router.get("/students/import/template")
def download_template(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("students.manage"))):
    """Download Excel template for student import"""
    import io
    try:
        import openpyxl
    except ImportError:
        return PlainTextResponse("openpyxl not installed", status_code=500)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Template"
    
    # Headers
    ws.append(["name", "email", "password", "roll_no", "branch_code", "semester_number", "phone", "guardian_name"])
    
    # Example row
    ws.append(["John Doe", "john@example.com", "Pass123@", "BCA001", "BCA", "1", "9876543210", "Robert Doe"])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=student_import_template.xlsx"}
    )


@router.get("/students/import")
def import_students_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("students.manage"))):
    """Show import students page"""
    return request.app.state.templates.TemplateResponse(
        "admin/import_students.html",
        {"request": request, "title": "Import Students"},
    )


@router.post("/students/import")
def import_students(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("students.manage"))):
    """Import students - placeholder for bulk import"""
    return request.app.state.templates.TemplateResponse(
        "admin/import_students.html",
        {"request": request, "message": "Bulk import feature coming soon. Use manual entry for now.", "title": "Import Students"},
    )
