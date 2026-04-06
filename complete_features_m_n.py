# Complete Features M and N - Export Excel and Bulk Import

print("=== Adding Export and Import Routes ===")

# Add export routes to admin.py
with open('app/routers/admin.py', 'r') as f:
    admin_content = f.read()

# Export Students to Excel
export_students_route = '''

# ==================== EXPORT TO EXCEL ====================

@router.get("/students/export")
def export_students(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("students.view"))):
    """Export students to Excel"""
    import io
    try:
        import openpyxl
    except ImportError:
        return PlainTextResponse("openpyxl not installed", status_code=500)
    
    students = db.query(models.StudentProfile).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    
    # Headers
    ws.append(["Roll No", "Name", "Email", "Branch", "Semester", "Phone", "Fee Status"])
    
    for student in students:
        user = db.query(models.User).filter(models.User.id == student.user_id).first()
        branch = db.query(models.Branch).filter(models.Branch.id == student.branch_id).first() if student.branch_id else None
        semester = db.query(models.Semester).filter(models.Semester.id == student.semester_id).first() if student.semester_id else None
        
        # Get fee status
        fee = db.query(models.FeeRecord).filter(models.FeeRecord.student_id == student.id).first()
        fee_status = fee.status if fee else "No Fee Record"
        
        ws.append([
            student.roll_no,
            user.name if user else "",
            user.email if user else "",
            branch.name if branch else "",
            semester.semester_number if semester else "",
            student.phone or "",
            fee_status
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=students.xlsx"}
    )
'''

# Export Fees to Excel
export_fees_route = '''

@router.get("/fees/export")
def export_fees(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("fees.view"))):
    """Export fees to Excel"""
    import io
    try:
        import openpyxl
    except ImportError:
        return PlainTextResponse("openpyxl not installed", status_code=500)
    
    fees = db.query(models.FeeRecord).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fees"
    
    # Headers
    ws.append(["Student Name", "Roll No", "Branch", "Semester", "Amount Due", "Amount Paid", "Pending", "Status"])
    
    for fee in fees:
        student = db.query(models.StudentProfile).filter(models.StudentProfile.id == fee.student_id).first()
        if student:
            user = db.query(models.User).filter(models.User.id == student.user_id).first()
            branch = db.query(models.Branch).filter(models.Branch.id == student.branch_id).first() if student.branch_id else None
            semester = db.query(models.Semester).filter(models.Semester.id == student.semester_id).first() if student.semester_id else None
            
            pending = fee.amount_due - fee.amount_paid
            
            ws.append([
                user.name if user else "",
                student.roll_no,
                branch.name if branch else "",
                semester.semester_number if semester else "",
                fee.amount_due,
                fee.amount_paid,
                pending,
                fee.status
            ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=fees.xlsx"}
    )
'''

# Export Attendance to Excel
export_attendance_route = '''

@router.get("/attendance/report/export")
def export_attendance_report(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("attendance.view"))):
    """Export attendance report to Excel"""
    import io
    try:
        import openpyxl
    except ImportError:
        return PlainTextResponse("openpyxl not installed", status_code=500)
    
    branch_id = request.query_params.get("branch_id")
    semester_id = request.query_params.get("semester_id")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Headers
    ws.append(["Roll Number", "Student Name", "Total Classes", "Present", "Absent", "Percentage"])
    
    if branch_id and semester_id:
        students = db.query(models.StudentProfile).filter(
            models.StudentProfile.branch_id == int(branch_id),
            models.StudentProfile.semester_id == int(semester_id)
        ).all()
        
        for student in students:
            user = db.query(models.User).filter(models.User.id == student.user_id).first()
            enrollments = db.query(models.Enrollment).filter(
                models.Enrollment.student_id == student.id
            ).all()
            
            total_classes = 0
            present = 0
            absent = 0
            
            for enrollment in enrollments:
                attendance = db.query(models.Attendance).filter(
                    models.Attendance.enrollment_id == enrollment.id
                ).all()
                total_classes += len(attendance)
                present += sum(1 for a in attendance if a.status == "present")
                absent += sum(1 for a in attendance if a.status == "absent")
            
            pct = round((present / total_classes * 100), 2) if total_classes > 0 else 0
            
            ws.append([
                student.roll_no,
                user.name if user else "",
                total_classes,
                present,
                absent,
                f"{pct}%"
            ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=attendance_report.xlsx"}
    )
'''

# Import Template Download
import_template_route = '''

# ==================== BULK STUDENT IMPORT ====================

@router.get("/students/import/template")
def download_template(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("students.manage"))):
    """Download Excel template for student import"""
    import io
    try:
        import openpyxl
    except ImportError:
        return PlainTextResponse("openpyxl not installed", status_code=500)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Template"
    
    # Headers
    ws.append(["name", "email", "password", "roll_no", "branch_code", "semester_number", "phone", "guardian_name"])
    
    # Example row
    ws.append(["John Doe", "john@example.com", "Pass123@", "BCA001", "BCA", "1", "9876543210", "Robert Doe"])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=student_import_template.xlsx"}
    )


@router.get("/students/import")
def import_students_page(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("students.manage"))):
    """Show import students page"""
    return request.app.state.templates.TemplateResponse(
        "admin/import_students.html",
        {"request": request, "title": "Import Students"},
    )


@router.post("/students/import")
def import_students(request: Request, db: Session = Depends(get_db), _=Depends(require_perm("students.manage"))):
    """Import students from Excel/CSV"""
    from fastapi import UploadFile, File
    from .auth import hash_password
    
    # Get the form data
    form = await request.form()
    file = form.get("file")
    
    if not file:
        return request.app.state.templates.TemplateResponse(
            "admin/import_students.html",
            {"request": request, "error": "No file uploaded", "title": "Import Students"},
        )
    
    success_count = 0
    failed = []
    
    try:
        # Read file content
        content = await file.read()
        
        # Try to read as Excel
        try:
            import openpyxl
            from io import BytesIO
            df = pd.read_excel(BytesIO(content))
        except:
            # Try CSV
            from io import StringIO
            df = pd.read_csv(StringIO(content.decode('utf-8')))
        
        # Process each row
        for index, row in df.iterrows():
            try:
                # Find branch
                branch = db.query(models.Branch).filter(
                    models.Branch.code == str(row.get('branch_code', ''))
                ).first()
                
                if not branch:
                    failed.append(f"Row {index+2}: Branch '{row.get('branch_code')}' not found")
                    continue
                
                # Find semester
                semester_num = int(row.get('semester_number', 1))
                semester = db.query(models.Semester).filter(
                    models.Semester.branch_id == branch.id,
                    models.Semester.semester_number == semester_num
                ).first()
                
                if not semester:
                    failed.append(f"Row {index+2}: Semester {semester_num} not found for branch {branch.name}")
                    continue
                
                # Check if email exists
                existing_user = db.query(models.User).filter(
                    models.User.email == str(row.get('email', '')).lower()
                ).first()
                
                if existing_user:
                    failed.append(f"Row {index+2}: Email '{row.get('email')}' already exists")
                    continue
                
                # Create user
                user = models.User(
                    name=str(row.get('name', '')),
                    email=str(row.get('email', '')).lower(),
                    password_hash=hash_password(str(row.get('password', 'student123'))),
                    role='student'
                )
                db.add(user)
                db.flush()
                
                # Create student profile
                profile = models.StudentProfile(
                    user_id=user.id,
                    roll_no=str(row.get('roll_no', '')),
                    department=branch.name,
                    branch_id=branch.id,
                    semester_id=semester.id,
                    phone=str(row.get('phone', '')),
                    guardian_name=str(row.get('guardian_name', ''))
                )
                db.add(profile)
                db.commit()
                
                success_count += 1
                
            except Exception as e:
                db.rollback()
                failed.append(f"Row {index+2}: {str(e)}")
    
    except Exception as e:
        return request.app.state.templates.TemplateResponse(
            "admin/import_students.html",
            {"request": request, "error": f"Error reading file: {str(e)}", "title": "Import Students"},
        )
    
    result_message = f"{success_count} students created successfully"
    if failed:
        result_message += f", {len(failed)} failed: " + "; ".join(failed[:5])
        if len(failed) > 5:
            result_message += f" ... and {len(failed) - 5} more"
    
    return request.app.state.templates.TemplateResponse(
        "admin/import_students.html",
        {"request": request, "message": result_message, "title": "Import Students"},
    )
'''

# Add routes to admin.py
admin_content = admin_content + export_students_route + export_fees_route + export_attendance_route + import_template_route

with open('app/routers/admin.py', 'w') as f:
    f.write(admin_content)

print("Export and Import routes added")
print("\n=== Features M and N Complete ===")
